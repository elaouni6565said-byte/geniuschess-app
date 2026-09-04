import pytest
import io
import openpyxl
from academy.models import Student
from portal.excel_export import export_students_to_excel

@pytest.mark.django_db
def test_export_excel_bilingual_utf8():
    """Tests Excel export with mixed French and Arabic characters without corruption (47.14 & 47.15)."""
    students = Student.objects.all()
    excel_bytes = export_students_to_excel(students, lang='ar')

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.title == "قائمة التلاميذ"
    assert ws.sheet_view.rightToLeft is True

    # Search for any student name in Arabic and French
    first_student = students.first()
    assert first_student is not None, "At least one student should exist"

    found_arabic = False
    found_french = False
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and first_student.first_name_ar in str(cell):
                found_arabic = True
            if cell and first_student.first_name_fr in str(cell):
                found_french = True

    assert found_arabic, f"Arabic student name ({first_student.first_name_ar}) should be present in Excel"
    assert found_french, f"French student name ({first_student.first_name_fr}) should be present in Excel"


@pytest.mark.django_db
def test_export_paid_and_unpaid_excel():
    """Tests Excel export for paid payments and unpaid invoices with French and Arabic."""
    from finance.models import Payment, Invoice
    from portal.excel_export import export_paid_payments_to_excel, export_unpaid_invoices_to_excel

    # 1. Paid payments export
    payments = Payment.objects.all()
    excel_paid_bytes = export_paid_payments_to_excel(payments, lang='fr')
    wb_paid = openpyxl.load_workbook(io.BytesIO(excel_paid_bytes))
    ws_paid = wb_paid.active
    assert "Paiements" in ws_paid.title
    # Check header exists
    headers = [cell for cell in next(ws_paid.iter_rows(min_row=3, max_row=3, values_only=True))]
    assert any("Reçu" in str(h) for h in headers if h)

    # 2. Unpaid invoices export in Arabic RTL
    invoices = Invoice.objects.filter(status__in=['unpaid', 'partial'])
    excel_unpaid_bytes = export_unpaid_invoices_to_excel(invoices, lang='ar')
    wb_unpaid = openpyxl.load_workbook(io.BytesIO(excel_unpaid_bytes))
    ws_unpaid = wb_unpaid.active
    assert ws_unpaid.sheet_view.rightToLeft is True
    assert "المستحقات" in ws_unpaid.title


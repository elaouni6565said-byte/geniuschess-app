import pytest
from decimal import Decimal
from datetime import date
from django.test import Client
import openpyxl
import io

from academy.models import User, Trainer, Subject, Group, SessionSchedule
from finance.models import TrainerPayout, Expense, ExpenseCategory
from finance.trainer_slip_pdf import generate_trainer_slip_pdf
from portal.excel_export import export_trainers_payroll_to_excel


@pytest.fixture
def admin_client():
    admin = User.objects.create_superuser(
        username="admin_trainer_test",
        password="adminpassword123",
        role="admin"
    )
    client = Client()
    client.force_login(admin)
    return client, admin


@pytest.mark.django_db
def test_trainer_model_and_bilingual_names():
    tr = Trainer.objects.create(
        first_name_fr="Karim",
        last_name_fr="Alami",
        first_name_ar="كريم",
        last_name_ar="العلمي",
        cin="G998877",
        phone="0661122334",
        specialty="Échecs / الشطرنج",
        compensation_type="per_session",
        default_rate=Decimal("150.00"),
        bank_name="Attijariwafa Bank",
        bank_rib="007780000123456789012345",
        active=True
    )

    assert tr.get_full_name("fr") == "Karim Alami"
    assert tr.get_full_name("ar") == "كريم العلمي"
    assert tr.get_bilingual_full_name() == "Karim Alami / كريم العلمي"
    assert "Karim Alami" in str(tr)


@pytest.mark.django_db
def test_trainer_payout_calculations():
    tr = Trainer.objects.create(
        first_name_fr="Youssef",
        last_name_fr="Bennani",
        cin="A123456",
        phone="0600000001",
        compensation_type="per_session",
        default_rate=Decimal("120.00")
    )

    payout = TrainerPayout.objects.create(
        trainer=tr,
        period_month=5,
        period_year=2026,
        compensation_type="per_session",
        sessions_count=10,
        rate_applied=Decimal("120.00"),
        bonus_amount=Decimal("200.00"),
        bonus_description="Prime tournoi",
        deduction_amount=Decimal("50.00"),
        deduction_description="Acompte",
        status="draft"
    )

    # Base = 10 * 120 = 1200 DH
    # Net = 1200 + 200 - 50 = 1350 DH
    assert payout.base_amount == Decimal("1200.00")
    assert payout.net_amount == Decimal("1350.00")
    assert payout.payout_number == f"BON-202605-{tr.id:02d}"


@pytest.mark.django_db
def test_payout_auto_sync_with_expense(admin_client):
    _, admin_user = admin_client
    tr = Trainer.objects.create(
        first_name_fr="Samir",
        last_name_fr="Idrissi",
        cin="B654321",
        phone="0600000002",
        compensation_type="monthly_fixed",
        default_rate=Decimal("2500.00")
    )

    payout = TrainerPayout.objects.create(
        trainer=tr,
        period_month=6,
        period_year=2026,
        compensation_type="monthly_fixed",
        base_amount=Decimal("2500.00"),
        rate_applied=Decimal("2500.00"),
        bonus_amount=Decimal("0.00"),
        deduction_amount=Decimal("0.00"),
        status="draft",
        payment_method="transfer",
        created_by=admin_user
    )

    # Au départ draft -> aucune dépense liée
    assert payout.expense is None

    # 1. Passer à paid -> crée automatiquement la dépense
    payout.status = "paid"
    payout.payment_date = date(2026, 6, 30)
    payout.save()
    payout.sync_with_expense()

    payout.refresh_from_db()
    assert payout.expense is not None
    exp = payout.expense
    assert exp.amount == Decimal("2500.00")
    assert exp.category.name_fr == "Rémunération & Honoraires Formateurs"
    assert exp.payment_method == "transfer"
    assert "Samir Idrissi" in exp.beneficiary
    assert exp.invoice_number == payout.payout_number

    # 2. Modifier la prime et resynchroniser
    payout.bonus_amount = Decimal("300.00")
    payout.bonus_description = "Encadrement championnat"
    payout.save()
    payout.sync_with_expense()

    payout.refresh_from_db()
    assert payout.net_amount == Decimal("2800.00")
    exp.refresh_from_db()
    assert exp.amount == Decimal("2800.00")

    # 3. Repasser à draft -> supprime la dépense
    payout.status = "draft"
    payout.save()
    payout.sync_with_expense()

    payout.refresh_from_db()
    assert payout.expense is None
    assert not Expense.objects.filter(invoice_number=payout.payout_number).exists()


@pytest.mark.django_db
def test_trainer_slip_pdf_generation():
    tr = Trainer.objects.create(
        first_name_fr="Hassan",
        last_name_fr="Mansouri",
        first_name_ar="حسن",
        last_name_ar="منصوري",
        cin="D112233",
        phone="0677889900",
        specialty="Échecs",
        compensation_type="per_session",
        default_rate=Decimal("150.00"),
        bank_name="CIH Bank",
        bank_rib="230780000987654321098765"
    )

    payout = TrainerPayout.objects.create(
        trainer=tr,
        period_month=7,
        period_year=2026,
        compensation_type="per_session",
        sessions_count=8,
        rate_applied=Decimal("150.00"),
        bonus_amount=Decimal("100.00"),
        deduction_amount=Decimal("0.00"),
        status="paid",
        payment_date=date(2026, 7, 28),
        payment_method="transfer"
    )

    # PDF French
    pdf_fr = generate_trainer_slip_pdf(payout, lang="fr")
    assert pdf_fr.startswith(b"%PDF-")
    assert len(pdf_fr) > 1000

    # PDF Arabic
    pdf_ar = generate_trainer_slip_pdf(payout, lang="ar")
    assert pdf_ar.startswith(b"%PDF-")
    assert len(pdf_ar) > 1000


@pytest.mark.django_db
def test_trainers_payroll_excel_export():
    tr = Trainer.objects.create(
        first_name_fr="Tarik",
        last_name_fr="Chraibi",
        cin="E554433",
        phone="0611223344",
        specialty="Calcul Mental",
        compensation_type="per_session",
        default_rate=Decimal("100.00")
    )

    TrainerPayout.objects.create(
        trainer=tr,
        period_month=5,
        period_year=2026,
        compensation_type="per_session",
        sessions_count=12,
        rate_applied=Decimal("100.00"),
        bonus_amount=Decimal("150.00"),
        deduction_amount=Decimal("50.00"),
        status="paid",
        payment_date=date(2026, 5, 30),
        payment_method="cash"
    )

    excel_bytes = export_trainers_payroll_to_excel(month=5, year=2026, lang="fr")
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    assert "Bordereau Honoraires" in wb.sheetnames
    ws = wb["Bordereau Honoraires"]
    assert "GENIUS CHESS ACADEMY" in ws["A1"].value
    assert "Sidi Kacem" in ws["A2"].value


@pytest.mark.django_db
def test_trainer_http_views_and_actions(admin_client):
    client, _ = admin_client

    # 1. Créer un formateur via POST
    tr_data = {
        "first_name_fr": "Nabil",
        "last_name_fr": "Rachidi",
        "first_name_ar": "نبيل",
        "last_name_ar": "رشيدي",
        "cin": "F998877",
        "phone": "0655443322",
        "email": "nabil@gca.ma",
        "specialty": "Robotique",
        "address": "Sidi Kacem",
        "compensation_type": "per_session",
        "default_rate": "150.00",
        "bank_name": "BMCE",
        "bank_rib": "011780000111222333444555",
        "active": True
    }
    resp = client.post("/trainers/add/", tr_data)
    assert resp.status_code == 302
    trainer = Trainer.objects.get(cin="F998877")
    assert trainer.last_name_fr == "Rachidi"

    # 2. Consulter la liste
    resp_list = client.get("/trainers/")
    assert resp_list.status_code == 200
    assert "Rachidi" in resp_list.content.decode("utf-8")

    # 3. Créer un bulletin via POST
    payout_data = {
        "trainer": trainer.id,
        "period_month": "8",
        "period_year": "2026",
        "compensation_type": "per_session",
        "sessions_count": "6",
        "rate_applied": "150.00",
        "base_amount": "900.00",
        "bonus_amount": "100.00",
        "bonus_description": "Prime stage",
        "deduction_amount": "0.00",
        "deduction_description": "",
        "status": "draft",
        "payment_date": "",
        "payment_method": "cash",
        "reference": "",
        "notes": "Test bulletin",
    }
    resp_pay = client.post("/trainers/payouts/add/", payout_data)
    assert resp_pay.status_code == 302
    payout = TrainerPayout.objects.get(trainer=trainer, period_month=8, period_year=2026)
    assert payout.net_amount == Decimal("1000.00")

    # 4. Fiche détaillée
    resp_detail = client.get(f"/trainers/payouts/{payout.id}/")
    assert resp_detail.status_code == 200
    assert "1000" in resp_detail.content.decode("utf-8")

    # 5. Marquer comme payé via POST
    resp_mark = client.post(f"/trainers/payouts/{payout.id}/mark-paid/")
    assert resp_mark.status_code == 302
    payout.refresh_from_db()
    assert payout.status == "paid"
    assert payout.expense is not None
    assert payout.expense.amount == Decimal("1000.00")

    # 6. Téléchargement PDF
    resp_pdf = client.get(f"/trainers/payouts/{payout.id}/pdf/?lang=fr")
    assert resp_pdf.status_code == 200
    assert resp_pdf["Content-Type"] == "application/pdf"

    # 7. Téléchargement Excel
    resp_excel = client.get("/trainers/payouts/excel/?month=8&year=2026&lang=fr")
    assert resp_excel.status_code == 200
    assert resp_excel["Content-Type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 8. API helper info
    resp_api = client.get(f"/api/trainers/{trainer.id}/info/?month=8&year=2026")
    assert resp_api.status_code == 200
    api_data = resp_api.json()
    assert api_data["id"] == trainer.id
    assert api_data["cin"] == "F998877"
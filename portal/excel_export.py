import io
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_FILL = PatternFill(start_color="001B57", end_color="001B57", fill_type="solid")
RED_FILL = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
GREEN_FILL = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
UNPAID_TOTAL_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="001B57")
TITLE_RED_FONT = Font(name="Segoe UI", size=14, bold=True, color="991B1B")
TOTAL_FONT = Font(name="Segoe UI", size=11, bold=True, color="001B57")
TOTAL_RED_FONT = Font(name="Segoe UI", size=11, bold=True, color="991B1B")
DATA_FONT = Font(name="Segoe UI", size=10)
BOLD_DATA_FONT = Font(name="Segoe UI", size=10, bold=True)

BORDER_THIN = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1'),
)

BORDER_TOTAL = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='double', color='001B57'),
)


def export_students_to_excel(students_queryset, lang="fr"):
    """
    Generates a comprehensive bilingual Excel workbook (.xlsx) of ALL students.
    Supports French, Arabic (RTL), and Bilingual.
    """
    wb = Workbook()
    ws = wb.active

    if lang == "ar":
        ws.title = "قائمة التلاميذ"
        ws.sheet_view.rightToLeft = True
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="right", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - لائحة جميع التلاميذ المسجلين 2026"
        headers = [
            "رقم التسجيل", "الاسم بالعربية", "الاسم بالفرنسية", "تاريخ الازدياد",
            "النشاط والمستوى", "المجموعة", "ولي الأمر", "الهاتف", "البريد الإلكتروني", "الحالة"
        ]
    elif lang == "bilingual":
        ws.title = "Élèves - التلاميذ"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste Complète des Élèves / لائحة التلاميذ 2026"
        headers = [
            "Matricule / التسجيل", "Nom (FR)", "الاسم (AR)", "Date Naissance",
            "Activité / النشاط", "Groupe / المجموعة", "Parent / ولي الأمر", "Tél / الهاتف", "Email", "Statut / الحالة"
        ]
    else: # fr
        ws.title = "Liste des Élèves"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste Complète des Élèves Inscrits 2026"
        headers = [
            "Matricule", "Nom (Français)", "Nom (Arabe)", "Date de Naissance",
            "Activité & Niveau", "Groupe", "Parent / Tuteur", "Téléphone", "Email", "Statut"
        ]

    # Title row
    last_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers row
    ws.row_dimensions[3].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = align_header
        cell.border = BORDER_THIN

    # Data rows
    row_idx = 4
    for st in students_queryset:
        ws.row_dimensions[row_idx].height = 22
        group = st.groups.first()
        subject_str = group.subject.get_bilingual_name() if group and group.subject else ("Échecs / الشطرنج" if st.active else "-")
        group_str = group.get_name(lang) if group else "-"
        parent_str = st.parent.get_name(lang) if st.parent else "-"
        phone_str = st.parent.phone if st.parent else "-"
        email_str = st.parent.email if (st.parent and st.parent.email) else "-"
        birth_str = st.birth_date.strftime("%d/%m/%Y") if st.birth_date else "-"
        status_str = "Actif / نشط" if st.active else "Inactif / غير نشط"

        row_values = [
            st.registration_number,
            f"{st.first_name_ar} {st.last_name_ar}".strip(),
            f"{st.first_name_fr} {st.last_name_fr}".strip(),
            birth_str,
            subject_str,
            group_str,
            parent_str,
            phone_str,
            email_str,
            status_str,
        ]
        if lang == "ar":
            row_values[1], row_values[2] = row_values[2], row_values[1]

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = DATA_FONT
            cell.alignment = align_data
            cell.border = BORDER_THIN
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def export_paid_payments_to_excel(payments_queryset, lang="fr"):
    """
    Generates an official Excel workbook (.xlsx) of all received / paid payments.
    Includes totals row, bilingual support, and formatted amounts.
    """
    wb = Workbook()
    ws = wb.active

    if lang == "ar":
        ws.title = "المقبوضات والأداءات"
        ws.sheet_view.rightToLeft = True
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="right", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - قائمة المقبوضات والأداءات المؤداة 2026"
        headers = [
            "رقم الإيصال", "تاريخ الأداء", "رقم التسجيل", "اسم التلميذ (بالعربية)", "اسم التلميذ (بالفرنسية)",
            "المادة / النشاط", "المجموعة", "ولي الأمر", "الهاتف", "طريقة الأداء", "المبلغ المؤدى (درهم)"
        ]
    elif lang == "bilingual":
        ws.title = "Paiements Réglés"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste des Paiements Réglés / قائمة المقبوضات 2026"
        headers = [
            "N° Reçu", "Date Paiement", "Matricule", "Élève (FR)", "الاسم (AR)",
            "Activité / النشاط", "Groupe", "Parent / ولي الأمر", "Téléphone", "Mode Règlement", "Montant (DH)"
        ]
    else: # fr
        ws.title = "Liste des Paiements"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste des Paiements Encaissés (Payants) 2026"
        headers = [
            "N° Reçu", "Date Paiement", "Matricule", "Nom Élève (FR)", "Nom Élève (AR)",
            "Activité", "Groupe", "Parent / Tuteur", "Téléphone", "Mode Règlement", "Montant Réglé (DH)"
        ]

    # Title row
    last_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers row
    ws.row_dimensions[3].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = align_header
        cell.border = BORDER_THIN

    # Data rows
    row_idx = 4
    total_amount = 0.0

    for p in payments_queryset:
        ws.row_dimensions[row_idx].height = 22
        st = p.student
        group = p.invoice.group if (p.invoice and p.invoice.group) else (st.groups.first() if st else None)
        subject_str = group.subject.get_name(lang) if (group and group.subject) else "-"
        group_str = group.get_name(lang) if group else "-"
        parent_str = st.parent.get_name(lang) if (st and st.parent) else "-"
        phone_str = st.parent.phone if (st and st.parent) else "-"
        
        # Payment method localized
        method_labels = {
            'cash': 'Espèces / نقداً',
            'bank_transfer': 'Virement bancaire / تحويل بنكي',
            'check': 'Chèque / شيك',
            'online': 'En ligne / أداء إلكتروني',
        }
        method_str = method_labels.get(p.payment_method, p.payment_method or "Espèces")
        amt = float(p.amount)
        total_amount += amt

        row_values = [
            f"#{p.receipt_number}",
            p.payment_date.strftime("%d/%m/%Y"),
            st.registration_number if st else "-",
            f"{st.first_name_fr} {st.last_name_fr}".strip() if st else "-",
            f"{st.first_name_ar} {st.last_name_ar}".strip() if st else "-",
            subject_str,
            group_str,
            parent_str,
            phone_str,
            method_str,
            amt,
        ]
        if lang == "ar":
            row_values[3], row_values[4] = row_values[4], row_values[3]

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = DATA_FONT
            cell.border = BORDER_THIN
            if col_num == len(headers):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
                cell.font = BOLD_DATA_FONT
            else:
                cell.alignment = align_data
        row_idx += 1

    # Total Summary Row
    ws.row_dimensions[row_idx].height = 28
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers) - 1)
    tot_label = ws.cell(row=row_idx, column=1)
    tot_label.value = "TOTAL DES RECETTES ENCAISSÉES / مجموع المقبوضات :" if lang != "ar" else "مجموع المبالغ المؤداة المحصلة :"
    tot_label.font = TOTAL_FONT
    tot_label.fill = TOTAL_FILL
    tot_label.alignment = Alignment(horizontal="right" if lang != "ar" else "left", vertical="center")
    tot_label.border = BORDER_TOTAL

    tot_val = ws.cell(row=row_idx, column=len(headers))
    tot_val.value = total_amount
    tot_val.font = TOTAL_FONT
    tot_val.fill = TOTAL_FILL
    tot_val.alignment = Alignment(horizontal="right", vertical="center")
    tot_val.number_format = '#,##0.00 "DH"'
    tot_val.border = BORDER_TOTAL

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def export_unpaid_invoices_to_excel(invoices_queryset, lang="fr"):
    """
    Generates an official Excel workbook (.xlsx) of all unpaid / partially paid students (Impayés).
    Includes contact info for reminders, outstanding balance, and total unpaid amount.
    """
    wb = Workbook()
    ws = wb.active

    if lang == "ar":
        ws.title = "المستحقات غير المؤداة"
        ws.sheet_view.rightToLeft = True
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="right", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - لائحة المستحقات غير المؤداة (المتأخرات) 2026"
        headers = [
            "رقم التسجيل", "اسم التلميذ (بالعربية)", "اسم التلميذ (بالفرنسية)", "ولي الأمر",
            "رقم الهاتف للمتابعة", "المادة / النشاط", "الشهر المعني", "الواجب الشهري (درهم)",
            "المبلغ المدفوع (درهم)", "الباقي المستحق (درهم)", "الحالة"
        ]
    elif lang == "bilingual":
        ws.title = "Impayés - المتأخرات"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste des Impayés / لائحة المستحقات غير المؤداة 2026"
        headers = [
            "Matricule", "Élève (FR)", "الاسم (AR)", "Parent / ولي الأمر",
            "Tél Relance", "Activité / النشاط", "Mois", "Montant Dû (DH)",
            "Payé (DH)", "Reste Impayé (DH)", "Statut / الحالة"
        ]
    else: # fr
        ws.title = "Liste des Impayés"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Liste des Élèves Non-Payants & Impayés 2026"
        headers = [
            "Matricule", "Nom Élève (FR)", "Nom Élève (AR)", "Parent / Tuteur",
            "Téléphone Relance", "Activité & Niveau", "Mois Concerné", "Montant Dû (DH)",
            "Déjà Versé (DH)", "Reste Impayé (DH)", "Statut"
        ]

    # Title row
    last_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = TITLE_RED_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers row
    ws.row_dimensions[3].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = RED_FILL
        cell.alignment = align_header
        cell.border = BORDER_THIN

    # Data rows
    row_idx = 4
    total_due = 0.0
    total_paid = 0.0
    total_balance = 0.0

    for inv in invoices_queryset:
        ws.row_dimensions[row_idx].height = 22
        st = inv.student
        group = inv.group if inv.group else (st.groups.first() if st else None)
        subject_str = group.subject.get_name(lang) if (group and group.subject) else "-"
        parent_str = st.parent.get_name(lang) if (st and st.parent) else "-"
        phone_str = st.parent.phone if (st and st.parent) else "-"
        month_str = inv.get_period_label(lang)

        amt_due = float(inv.amount_due)
        amt_paid = float(inv.amount_paid)
        balance = float(inv.get_balance())

        total_due += amt_due
        total_paid += amt_paid
        total_balance += balance

        status_str = "Impayé / غير مؤدى" if inv.status == 'unpaid' else "Partiel / أداء جزئي"

        row_values = [
            st.registration_number if st else "-",
            f"{st.first_name_fr} {st.last_name_fr}".strip() if st else "-",
            f"{st.first_name_ar} {st.last_name_ar}".strip() if st else "-",
            parent_str,
            phone_str,
            subject_str,
            month_str,
            amt_due,
            amt_paid,
            balance,
            status_str,
        ]
        if lang == "ar":
            row_values[1], row_values[2] = row_values[2], row_values[1]

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = DATA_FONT
            cell.border = BORDER_THIN
            if col_num in (8, 9):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
            elif col_num == 10:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
            else:
                cell.alignment = align_data
        row_idx += 1

    # Total Summary Row
    ws.row_dimensions[row_idx].height = 28
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
    tot_label = ws.cell(row=row_idx, column=1)
    tot_label.value = "TOTAL DES IMPAYÉS RESTANTS / مجموع المتأخرات المتبقية :" if lang != "ar" else "مجموع المتأخرات المتبقية المستحقة :"
    tot_label.font = TOTAL_RED_FONT
    tot_label.fill = UNPAID_TOTAL_FILL
    tot_label.alignment = Alignment(horizontal="right" if lang != "ar" else "left", vertical="center")
    tot_label.border = BORDER_TOTAL

    tot_due_cell = ws.cell(row=row_idx, column=8)
    tot_due_cell.value = total_due
    tot_due_cell.font = BOLD_DATA_FONT
    tot_due_cell.fill = UNPAID_TOTAL_FILL
    tot_due_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_due_cell.number_format = '#,##0.00 "DH"'
    tot_due_cell.border = BORDER_TOTAL

    tot_paid_cell = ws.cell(row=row_idx, column=9)
    tot_paid_cell.value = total_paid
    tot_paid_cell.font = BOLD_DATA_FONT
    tot_paid_cell.fill = UNPAID_TOTAL_FILL
    tot_paid_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_paid_cell.number_format = '#,##0.00 "DH"'
    tot_paid_cell.border = BORDER_TOTAL

    tot_bal_cell = ws.cell(row=row_idx, column=10)
    tot_bal_cell.value = total_balance
    tot_bal_cell.font = TOTAL_RED_FONT
    tot_bal_cell.fill = UNPAID_TOTAL_FILL
    tot_bal_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_bal_cell.number_format = '#,##0.00 "DH"'
    tot_bal_cell.border = BORDER_TOTAL

    ws.cell(row=row_idx, column=11).border = BORDER_TOTAL
    ws.cell(row=row_idx, column=11).fill = UNPAID_TOTAL_FILL

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def export_expenses_to_excel(expenses_queryset, lang="fr", month=None, year=None):
    """
    Génère un classeur Excel officiel (.xlsx) du registre des dépenses et charges.
    Prend en charge le français, l'arabe (RTL) et le bilingue, avec ligne de totalisation.
    """
    wb = Workbook()
    ws = wb.active

    period_str = ""
    if month and year:
        period_str = f" - {month:02d}/{year}"
    elif year:
        period_str = f" - Année {year}"

    if lang == "ar":
        ws.title = "سجل المصاريف"
        ws.sheet_view.rightToLeft = True
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="right", vertical="center")
        title_text = f"GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - سجل المصاريف والنفقات{period_str}"
        headers = [
            "التاريخ", "بيان المصروف", "الصنف / الفئة", "المستفيد / الجهة",
            "رقم الفاتورة / الوصل", "طريقة الأداء", "المبلغ (درهم)", "ملاحظات"
        ]
    elif lang == "bilingual":
        ws.title = "Dépenses - المصاريف"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = f"GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Registre des Dépenses{period_str}"
        headers = [
            "Date / التاريخ", "Libellé / البيان", "Catégorie / الصنف", "Bénéficiaire / المستفيد",
            "N° Facture / الوصل", "Mode Règlement", "Montant / المبلغ (DH)", "Remarques / ملاحظات"
        ]
    else: # fr
        ws.title = "Dépenses"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = f"GENIUS CHESS ACADEMY - جمعية الشطرنج القاسمي - Registre des Dépenses et Charges{period_str}"
        headers = [
            "Date", "Libellé de la Dépense", "Catégorie", "Bénéficiaire / Fournisseur",
            "N° Facture / Réf", "Mode de Règlement", "Montant (DH)", "Observations"
        ]

    # Title row
    last_col_letter = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Subtitle / Coordinates row
    ws.merge_cells(f"A2:{last_col_letter}2")
    sub_cell = ws["A2"]
    sub_cell.value = "Sidi Kacem • www.geniuschess.ma • Tél: 06 060424142"
    sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Headers row
    ws.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = align_header
        cell.border = BORDER_THIN

    # Data rows
    row_idx = 5
    total_amount = 0.0

    for exp in expenses_queryset:
        ws.row_dimensions[row_idx].height = 22
        cat_name = exp.category.get_name(lang) if exp.category else "-"
        date_str = exp.expense_date.strftime("%d/%m/%Y") if exp.expense_date else "-"
        method_str = exp.get_method_label(lang)
        amt = float(exp.amount)
        total_amount += amt

        row_values = [
            date_str,
            exp.title,
            cat_name,
            exp.beneficiary or "-",
            exp.invoice_number or "-",
            method_str,
            amt,
            exp.notes or "-",
        ]

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = DATA_FONT
            cell.border = BORDER_THIN
            if col_num == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
                cell.font = BOLD_DATA_FONT
            elif col_num == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = align_data
        row_idx += 1

    # Total Summary Row
    ws.row_dimensions[row_idx].height = 28
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    tot_label = ws.cell(row=row_idx, column=1)
    tot_label.value = "TOTAL GÉNÉRAL DES DÉPENSES / مجموع المصاريف :" if lang != "ar" else "مجموع المصاريف الإجمالي :"
    tot_label.font = TOTAL_FONT
    tot_label.fill = TOTAL_FILL
    tot_label.alignment = Alignment(horizontal="right" if lang != "ar" else "left", vertical="center")
    tot_label.border = BORDER_TOTAL

    tot_cell = ws.cell(row=row_idx, column=7)
    tot_cell.value = total_amount
    tot_cell.font = TOTAL_FONT
    tot_cell.fill = TOTAL_FILL
    tot_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_cell.number_format = '#,##0.00 "DH"'
    tot_cell.border = BORDER_TOTAL

    notes_total_cell = ws.cell(row=row_idx, column=8)
    notes_total_cell.border = BORDER_TOTAL
    notes_total_cell.fill = TOTAL_FILL

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes


def export_annual_financial_report_to_excel(year, lang="fr", closing=None):
    """
    Génère un classeur Excel complet multi-feuilles (.xlsx) du Bilan Financier Annuel
    pour l'Assemblée Générale de l'Association جمعية الشطرنج القاسمي et Genius Chess Academy.
    Feuille 1: Synthèse du Bilan & Trésorerie
    Feuille 2: Détail des Recettes Encaissées
    Feuille 3: Détail des Dépenses Acquittées
    Feuille 4: Journal Chronologique de Trésorerie (Caisse & Banque)
    """
    from finance.models import Payment, Expense, ExpenseCategory
    from academy.models import Subject
    from django.db.models import Sum

    wb = Workbook()

    # Querysets
    payments_qs = Payment.objects.filter(payment_date__year=year).select_related('student', 'invoice', 'invoice__group').order_by('payment_date', 'id')
    expenses_qs = Expense.objects.filter(expense_date__year=year).select_related('category').order_by('expense_date', 'id')

    tot_collected = payments_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
    tot_expense = expenses_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
    net_result = tot_collected - tot_expense

    cash_rec = payments_qs.filter(payment_method='cash').aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
    bank_rec = payments_qs.filter(payment_method__in=['transfer', 'check', 'online']).aggregate(t=Sum('amount'))['t'] or Decimal('0.00')

    cash_exp = expenses_qs.filter(payment_method='cash').aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
    bank_exp = expenses_qs.filter(payment_method__in=['transfer', 'check']).aggregate(t=Sum('amount'))['t'] or Decimal('0.00')

    init_cash = closing.initial_cash_balance if closing else Decimal('0.00')
    init_bank = closing.initial_bank_balance if closing else Decimal('0.00')

    theo_cash = init_cash + cash_rec - cash_exp
    theo_bank = init_bank + bank_rec - bank_exp

    # =========================================================================
    # FEUILLE 1 : BILAN SYNTHÉTIQUE & RAPPORT AG
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Bilan Financier AG" if lang != "ar" else "التقرير المالي للجمع العام"
    if lang == "ar":
        ws1.sheet_view.rightToLeft = True

    title_text = f"GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي — RAPPORT FINANCIER EXERCICE {year}"
    sub_text = "Sidi Kacem • www.geniuschess.ma • Tél: 06 060424142"

    ws1.merge_cells("A1:G1")
    ws1["A1"] = title_text
    ws1["A1"].font = TITLE_FONT
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:G2")
    ws1["A2"] = sub_text
    ws1["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    # Section 1: KPI Summary
    ws1["A4"] = "1. SYNTHÈSE BUDGÉTAIRE GÉNÉRALE DE L'EXERCICE"
    ws1["A4"].font = Font(name="Segoe UI", size=11, bold=True, color="001B57")

    summary_headers = ["Poste Budgétaire", "Espèces (Caisse)", "Banque (Chèque/Virement)", "Total Général (DH)", "Appréciation"]
    ws1.row_dimensions[5].height = 24
    for col_idx, h_text in enumerate(summary_headers, 1):
        c = ws1.cell(row=5, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    summary_rows = [
        ("TOTAL DES RECETTES ENCAISSÉES (+)", float(cash_rec), float(bank_rec), float(tot_collected), "Cotisations & droits encaissés"),
        ("TOTAL DES DÉPENSES ACQUITTÉES (-)", float(cash_exp), float(bank_exp), float(tot_expense), "Coûts et charges d'exploitation"),
        ("RÉSULTAT NET D'EXPLOITATION", float(cash_rec - cash_exp), float(bank_rec - bank_exp), float(net_result), "Excédent d'exercice" if net_result >= 0 else "Déficit temporaire"),
    ]

    for r_idx, row_data in enumerate(summary_rows, 6):
        ws1.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BOLD_DATA_FONT if r_idx == 8 else DATA_FONT
            cell.border = BORDER_THIN
            if c_idx in (2, 3, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
                if r_idx == 8:
                    cell.font = TOTAL_FONT if net_result >= 0 else TOTAL_RED_FONT
                    cell.fill = TOTAL_FILL if net_result >= 0 else UNPAID_TOTAL_FILL
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Section 2: Trésorerie & Disponibilités
    ws1["A10"] = "2. ÉTAT DE TRÉSORERIE ET DISPONIBILITÉS LIQUIDES"
    ws1["A10"].font = Font(name="Segoe UI", size=11, bold=True, color="001B57")

    tres_headers = ["Trésorerie", "Solde Initial", "Flux Entrées (+)", "Flux Sorties (-)", "Solde Théorique", "Solde Réel Constaté", "Écart Constaté"]
    ws1.row_dimensions[11].height = 24
    for col_idx, h_text in enumerate(tres_headers, 1):
        c = ws1.cell(row=11, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    phys_cash_val = float(closing.physical_cash_counted) if (closing and closing.physical_cash_counted is not None) else float(theo_cash)
    bank_stmt_val = float(closing.bank_statement_balance) if (closing and closing.bank_statement_balance is not None) else float(theo_bank)

    tres_rows = [
        ("Caisse Espèces (Coffre)", float(init_cash), float(cash_rec), float(cash_exp), float(theo_cash), phys_cash_val, float(closing.cash_discrepancy if closing else 0)),
        ("Compte Bancaire", float(init_bank), float(bank_rec), float(bank_exp), float(theo_bank), bank_stmt_val, float(closing.bank_discrepancy if closing else 0)),
        ("TOTAL DISPONIBILITÉS", float(init_cash + init_bank), float(tot_collected), float(tot_expense), float(theo_cash + theo_bank), phys_cash_val + bank_stmt_val, float((closing.cash_discrepancy + closing.bank_discrepancy) if closing else 0)),
    ]

    for r_idx, row_data in enumerate(tres_rows, 12):
        ws1.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BOLD_DATA_FONT if r_idx == 14 else DATA_FONT
            cell.border = BORDER_TOTAL if r_idx == 14 else BORDER_THIN
            if c_idx >= 2:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '#,##0.00 "DH"'
                if r_idx == 14:
                    cell.fill = TOTAL_FILL
                    cell.font = TOTAL_FONT
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Section 3: Ventilation Dépenses par Catégorie
    ws1["A16"] = "3. VENTILATION DES CHARGES D'EXPLOITATION PAR CATÉGORIE"
    ws1["A16"].font = Font(name="Segoe UI", size=11, bold=True, color="001B57")

    exp_cat_headers = ["Catégorie de Dépense", "Nombre d'opérations", "Part Budgétaire (%)", "Montant Total (DH)"]
    ws1.row_dimensions[17].height = 24
    for col_idx, h_text in enumerate(exp_cat_headers, 1):
        c = ws1.cell(row=17, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    curr_r = 18
    tot_exp_float = float(tot_expense) if tot_expense > 0 else 1.0
    for cat in ExpenseCategory.objects.all():
        c_exps = expenses_qs.filter(category=cat)
        c_amt = c_exps.aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
        c_cnt = c_exps.count()
        if c_amt > 0 or c_cnt > 0:
            ws1.row_dimensions[curr_r].height = 20
            pct = round((float(c_amt) / tot_exp_float) * 100, 1)
            ws1.cell(row=curr_r, column=1, value=f"{cat.icon} {cat.name_fr} ({cat.name_ar})").alignment = Alignment(horizontal="left", vertical="center")
            ws1.cell(row=curr_r, column=2, value=c_cnt).alignment = Alignment(horizontal="center", vertical="center")
            ws1.cell(row=curr_r, column=3, value=f"{pct}%").alignment = Alignment(horizontal="right", vertical="center")
            amt_c = ws1.cell(row=curr_r, column=4, value=float(c_amt))
            amt_c.alignment = Alignment(horizontal="right", vertical="center")
            amt_c.number_format = '#,##0.00 "DH"'
            amt_c.font = BOLD_DATA_FONT
            for col_i in range(1, 5):
                ws1.cell(row=curr_r, column=col_i).border = BORDER_THIN
            curr_r += 1

    # Total Ligne Dépenses
    ws1.row_dimensions[curr_r].height = 24
    ws1.cell(row=curr_r, column=1, value="TOTAL DES CHARGES D'EXPLOITATION").font = TOTAL_FONT
    ws1.cell(row=curr_r, column=1).border = BORDER_TOTAL
    ws1.cell(row=curr_r, column=2, value=expenses_qs.count()).border = BORDER_TOTAL
    ws1.cell(row=curr_r, column=3, value="100.0%").border = BORDER_TOTAL
    tot_exp_cell = ws1.cell(row=curr_r, column=4, value=float(tot_expense))
    tot_exp_cell.font = TOTAL_FONT
    tot_exp_cell.fill = TOTAL_FILL
    tot_exp_cell.border = BORDER_TOTAL
    tot_exp_cell.alignment = Alignment(horizontal="right", vertical="center")
    tot_exp_cell.number_format = '#,##0.00 "DH"'

    # Signatures Footer
    curr_r += 3
    ws1.cell(row=curr_r, column=2, value="LE TRÉSORIER GÉNÉRAL").font = TOTAL_FONT
    ws1.cell(row=curr_r, column=5, value="LE PRÉSIDENT DE L'ASSOCIATION").font = TOTAL_FONT
    curr_r += 1
    ws1.cell(row=curr_r, column=2, value="أمين المال العام").font = BOLD_DATA_FONT
    ws1.cell(row=curr_r, column=5, value="رئيس الجمعية").font = BOLD_DATA_FONT

    # =========================================================================
    # FEUILLE 2 : DÉTAIL DES RECETTES ENCAISSÉES
    # =========================================================================
    ws2 = wb.create_sheet(title="Recettes Encaissées" if lang != "ar" else "المقبوضات")
    if lang == "ar":
        ws2.sheet_view.rightToLeft = True

    ws2.merge_cells("A1:H1")
    ws2["A1"] = f"GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي — Registre Détaillé des Recettes {year}"
    ws2["A1"].font = TITLE_FONT
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    rec_headers = ["N° Reçu", "Date Paiement", "Matricule", "Nom Élève (FR)", "Nom Élève (AR)", "Activité", "Mode Règlement", "Montant Réglé (DH)"]
    ws2.row_dimensions[3].height = 26
    for col_idx, h_text in enumerate(rec_headers, 1):
        c = ws2.cell(row=3, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    row_i = 4
    for p in payments_qs:
        ws2.row_dimensions[row_i].height = 20
        st = p.student
        grp = p.invoice.group if (p.invoice and p.invoice.group) else (st.groups.first() if st else None)
        sub_name = grp.subject.get_bilingual_name() if (grp and grp.subject) else "-"

        vals = [
            f"#{p.receipt_number}",
            p.payment_date.strftime("%d/%m/%Y"),
            st.registration_number if st else "-",
            f"{st.first_name_fr} {st.last_name_fr}".strip() if st else "-",
            f"{st.first_name_ar} {st.last_name_ar}".strip() if st else "-",
            sub_name,
            p.get_method_label(lang),
            float(p.amount)
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws2.cell(row=row_i, column=col_idx, value=val)
            c.font = DATA_FONT
            c.border = BORDER_THIN
            if col_idx == 8:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00 "DH"'
                c.font = BOLD_DATA_FONT
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        row_i += 1

    # Total Recettes
    ws2.row_dimensions[row_i].height = 26
    ws2.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=7)
    tot_rec_lbl = ws2.cell(row=row_i, column=1, value="TOTAL DES RECETTES ENCAISSÉES / مجموع المداخيل :")
    tot_rec_lbl.font = TOTAL_FONT
    tot_rec_lbl.fill = TOTAL_FILL
    tot_rec_lbl.border = BORDER_TOTAL
    tot_rec_lbl.alignment = Alignment(horizontal="right", vertical="center")

    tot_rec_val = ws2.cell(row=row_i, column=8, value=float(tot_collected))
    tot_rec_val.font = TOTAL_FONT
    tot_rec_val.fill = TOTAL_FILL
    tot_rec_val.border = BORDER_TOTAL
    tot_rec_val.alignment = Alignment(horizontal="right", vertical="center")
    tot_rec_val.number_format = '#,##0.00 "DH"'

    # =========================================================================
    # FEUILLE 3 : DÉTAIL DES DÉPENSES ACQUITTÉES
    # =========================================================================
    ws3 = wb.create_sheet(title="Dépenses Acquittées" if lang != "ar" else "المصاريف")
    if lang == "ar":
        ws3.sheet_view.rightToLeft = True

    ws3.merge_cells("A1:H1")
    ws3["A1"] = f"GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي — Registre Détaillé des Dépenses {year}"
    ws3["A1"].font = TITLE_FONT
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    exp_headers = ["Date", "Libellé de la Dépense", "Catégorie", "Bénéficiaire / Fournisseur", "N° Facture / Réf", "Mode Règlement", "Montant (DH)", "Observations"]
    ws3.row_dimensions[3].height = 26
    for col_idx, h_text in enumerate(exp_headers, 1):
        c = ws3.cell(row=3, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    row_i = 4
    for exp in expenses_qs:
        ws3.row_dimensions[row_i].height = 20
        cat_name = exp.category.get_name(lang) if exp.category else "-"
        vals = [
            exp.expense_date.strftime("%d/%m/%Y"),
            exp.title,
            cat_name,
            exp.beneficiary or "-",
            exp.invoice_number or "-",
            exp.get_method_label(lang),
            float(exp.amount),
            exp.notes or "-"
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws3.cell(row=row_i, column=col_idx, value=val)
            c.font = DATA_FONT
            c.border = BORDER_THIN
            if col_idx == 7:
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = '#,##0.00 "DH"'
                c.font = BOLD_DATA_FONT
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        row_i += 1

    # Total Dépenses
    ws3.row_dimensions[row_i].height = 26
    ws3.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=6)
    tot_exp_lbl = ws3.cell(row=row_i, column=1, value="TOTAL DES DÉPENSES ACQUITTÉES / مجموع المصاريف :")
    tot_exp_lbl.font = TOTAL_FONT
    tot_exp_lbl.fill = TOTAL_FILL
    tot_exp_lbl.border = BORDER_TOTAL
    tot_exp_lbl.alignment = Alignment(horizontal="right", vertical="center")

    tot_exp_val = ws3.cell(row=row_i, column=7, value=float(tot_expense))
    tot_exp_val.font = TOTAL_FONT
    tot_exp_val.fill = TOTAL_FILL
    tot_exp_val.border = BORDER_TOTAL
    tot_exp_val.alignment = Alignment(horizontal="right", vertical="center")
    tot_exp_val.number_format = '#,##0.00 "DH"'
    ws3.cell(row=row_i, column=8).border = BORDER_TOTAL
    ws3.cell(row=row_i, column=8).fill = TOTAL_FILL

    # =========================================================================
    # FEUILLE 4 : JOURNAL CHRONOLOGIQUE DE TRÉSORERIE (CAISSE & BANQUE)
    # =========================================================================
    ws4 = wb.create_sheet(title="Journal Trésorerie" if lang != "ar" else "سجل الخزينة")
    if lang == "ar":
        ws4.sheet_view.rightToLeft = True

    ws4.merge_cells("A1:G1")
    ws4["A1"] = f"GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي — Journal Chronologique des Flux de Trésorerie {year}"
    ws4["A1"].font = TITLE_FONT
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    jour_headers = ["Date", "N° Pièce / Réf", "Nature & Description de l'Opération", "Compte / Caisse", "Entrée (+) DH", "Sortie (-) DH", "Solde Cumulé (DH)"]
    ws4.row_dimensions[3].height = 26
    for col_idx, h_text in enumerate(jour_headers, 1):
        c = ws4.cell(row=3, column=col_idx, value=h_text)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN

    # Build chronological timeline
    timeline = []
    for p in payments_qs:
        acc_type = "Caisse Espèces" if p.payment_method == 'cash' else "Banque"
        timeline.append({
            'date': p.payment_date,
            'ref': f"#{p.receipt_number}",
            'desc': f"Cotisation - {p.student.get_full_name('fr')}",
            'account': acc_type,
            'in': float(p.amount),
            'out': 0.0,
        })

    for exp in expenses_qs:
        acc_type = "Caisse Espèces" if exp.payment_method == 'cash' else "Banque"
        timeline.append({
            'date': exp.expense_date,
            'ref': exp.invoice_number or f"DEP-{exp.id}",
            'desc': f"Dépense - {exp.title}",
            'account': acc_type,
            'in': 0.0,
            'out': float(exp.amount),
        })

    timeline.sort(key=lambda x: (x['date'], x['ref']))

    row_i = 4
    cumul = float(init_cash + init_bank)

    # Initial Balance Row
    ws4.cell(row=row_i, column=1, value=f"01/01/{year}").border = BORDER_THIN
    ws4.cell(row=row_i, column=2, value="REPORT").border = BORDER_THIN
    ws4.cell(row=row_i, column=3, value="SOLDE INITIAL REPORT À NOUVEAU").border = BORDER_THIN
    ws4.cell(row=row_i, column=4, value="Caisse & Banque").border = BORDER_THIN
    ws4.cell(row=row_i, column=5, value="-").border = BORDER_THIN
    ws4.cell(row=row_i, column=6, value="-").border = BORDER_THIN
    init_cumul_c = ws4.cell(row=row_i, column=7, value=cumul)
    init_cumul_c.border = BORDER_THIN
    init_cumul_c.number_format = '#,##0.00 "DH"'
    init_cumul_c.font = BOLD_DATA_FONT
    init_cumul_c.alignment = Alignment(horizontal="right", vertical="center")
    row_i += 1

    for item in timeline:
        ws4.row_dimensions[row_i].height = 20
        cumul += (item['in'] - item['out'])
        vals = [
            item['date'].strftime("%d/%m/%Y"),
            item['ref'],
            item['desc'],
            item['account'],
            item['in'] if item['in'] > 0 else "-",
            item['out'] if item['out'] > 0 else "-",
            cumul
        ]
        for col_idx, val in enumerate(vals, 1):
            c = ws4.cell(row=row_i, column=col_idx, value=val)
            c.font = DATA_FONT
            c.border = BORDER_THIN
            if col_idx in (5, 6, 7):
                c.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0.00 "DH"'
                if col_idx == 7:
                    c.font = BOLD_DATA_FONT
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        row_i += 1

    # Auto-adjust column widths for all sheets
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes



def export_trainers_payroll_to_excel(month, year, lang="fr"):
    """
    Génère un classeur Excel contenant le Bordereau Récapitulatif Mensuel
    des honoraires et indemnités des formateurs de l'académie et de l'association.
    """
    from finance.models import TrainerPayout
    from core.i18n import FRENCH_MONTHS, ARABIC_MONTHS
    from django.db.models import Sum

    wb = Workbook()
    ws = wb.active
    ws.title = "Bordereau Honoraires" if lang != "ar" else "بيان أجور ومستحقات المدربين"
    if lang == "ar":
        ws.sheet_view.rightToLeft = True

    payouts = TrainerPayout.objects.filter(
        period_month=month, period_year=year
    ).select_related('trainer').order_by('trainer__last_name_fr', 'trainer__first_name_fr')

    month_name = FRENCH_MONTHS.get(month, str(month)).capitalize() if lang != "ar" else ARABIC_MONTHS.get(month, str(month))
    title_text = f"GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي — ÉTAT DES HONORAIRES ({month_name} {year})"
    sub_text = "Sidi Kacem • www.geniuschess.ma • Tél: 06 060424142"

    ws.merge_cells("A1:N1")
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:N2")
    ws["A2"] = sub_text
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    headers = [
        "N° Bulletin",
        "Formateur (Bénéficiaire)",
        "CIN",
        "Discipline",
        "Mode Calcul",
        "Volume (Séances/H)",
        "Tarif Unitaire",
        "Base (DH)",
        "Primes (DH)",
        "Retenues (DH)",
        "Net à Payer (DH)",
        "Mode Paiement",
        "Date Versement",
        "Statut"
    ]
    if lang == "ar":
        headers = [
            "رقم البيان",
            "اسم المدرب (المستفيد)",
            "ر.ب.و (CIN)",
            "التخصص",
            "طريقة الاحتساب",
            "عدد الحصص",
            "التعريفة",
            "المبلغ الأساسي",
            "المنح والتعويضات",
            "الاقتطاعات والتسبيقات",
            "الصافي للأداء",
            "طريقة الأداء",
            "تاريخ الأداء",
            "الوضعية"
        ]

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER_THIN
    ws.row_dimensions[4].height = 26

    row_idx = 5
    tot_base = Decimal("0.00")
    tot_bonus = Decimal("0.00")
    tot_deduct = Decimal("0.00")
    tot_net = Decimal("0.00")

    comp_labels = {
        'monthly_fixed': 'Forfait mensuel',
        'per_session': 'Par séance',
        'per_hour': 'Horaire',
    }

    row_even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    row_odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for p in payouts:
        fill_to_use = row_even_fill if row_idx % 2 == 0 else row_odd_fill
        tr = p.trainer
        c_type = comp_labels.get(p.compensation_type, p.compensation_type)
        pay_date_val = p.payment_date.strftime("%d/%m/%Y") if p.payment_date else "---"

        row_vals = [
            p.payout_number,
            tr.get_bilingual_full_name() if lang != "ar" else tr.get_full_name("ar"),
            tr.cin or "---",
            tr.specialty or "Échecs",
            c_type,
            p.sessions_count,
            float(p.rate_applied),
            float(p.base_amount),
            float(p.bonus_amount),
            float(p.deduction_amount),
            float(p.net_amount),
            p.get_method_label(lang),
            pay_date_val,
            p.get_status_label(lang),
        ]

        tot_base += p.base_amount
        tot_bonus += p.bonus_amount
        tot_deduct += p.deduction_amount
        tot_net += p.net_amount

        for col_idx, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = DATA_FONT
            c.fill = fill_to_use
            c.border = BORDER_THIN
            if col_idx in [7, 8, 9, 10, 11]:
                c.number_format = '#,##0.00 "DH"'
                c.alignment = Alignment(horizontal="right", vertical="center")
                if col_idx == 11:
                    c.font = BOLD_DATA_FONT
            elif col_idx in [1, 3, 5, 6, 12, 13, 14]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

    # Ligne de Total
    ws.cell(row=row_idx, column=1, value="TOTAL DES HONORAIRES DU MOIS").font = BOLD_DATA_FONT
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)

    ws.cell(row=row_idx, column=8, value=float(tot_base)).number_format = '#,##0.00 "DH"'
    ws.cell(row=row_idx, column=9, value=float(tot_bonus)).number_format = '#,##0.00 "DH"'
    ws.cell(row=row_idx, column=10, value=float(tot_deduct)).number_format = '#,##0.00 "DH"'
    ws.cell(row=row_idx, column=11, value=float(tot_net)).number_format = '#,##0.00 "DH"'

    total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for c_i in range(1, 15):
        cell_tot = ws.cell(row=row_idx, column=c_i)
        cell_tot.font = BOLD_DATA_FONT
        cell_tot.border = BORDER_TOTAL
        cell_tot.fill = total_fill
        if c_i in [8, 9, 10, 11]:
            cell_tot.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[row_idx].height = 25

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes

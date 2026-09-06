import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles et palettes
NAVY_HEADER = "0A192F"
GOLD = "D97706"
GRAY_LIGHT = "F8FAFC"
BORDER_COLOR = "CBD5E1"

CLR_PRESENT = "DCFCE7"   # Vert clair
TXT_PRESENT = "166534"
CLR_ABSENT = "FEE2E2"    # Rouge clair
TXT_ABSENT = "991B1B"
CLR_LATE = "FFEDD5"      # Orange clair
TXT_LATE = "9A3412"
CLR_JUSTIFIED = "E0E7FF" # Bleu/Indigo clair
TXT_JUSTIFIED = "3730A3"


def generate_attendance_excel(attendances, student_summaries, period_label, lang="fr"):
    """
    Génère un classeur Excel complet (.xlsx) contenant :
    1. Onglet 'Émargements & Séances' : Liste détaillée classée par Activité, Groupe, Date et Heure.
    2. Onglet 'Synthèse par Élève' : Taux d'assiduité, présences, retards et absences par élève.
    """
    wb = Workbook()

    # Définition des bordures fines
    thin_side = Side(style="thin", color=BORDER_COLOR)
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # =========================================================================
    # FEUILLE 1 : DÉTAIL DES ÉMARGEMENTS
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Émargements par Séance" if lang == "fr" else "سجل الحضور بالحصص"
    ws1.views.sheetView[0].showGridLines = True

    # 1. En-tête principal
    ws1.merge_cells("A1:J1")
    title_cell = ws1["A1"]
    title_cell.value = "GENIUS CHESS ACADEMY — جمعية الشطرنج القاسمي"
    title_cell.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:J2")
    sub_cell = ws1["A2"]
    sub_cell.value = (
        f"Registre Récapitulatif des Présences • Période : {period_label} • Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        if lang == "fr" else
        f"السجل العام للحضور والغياب • الفترة : {period_label} • استخرج في {datetime.now().strftime('%d/%m/%Y على %H:%M')}"
    )
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
    sub_cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 22

    # Ligne vide
    ws1.row_dimensions[3].height = 10

    # 2. En-têtes de colonnes
    headers = [
        ("N°", 6),
        ("Activité / النشاط", 20),
        ("Groupe / الفوج", 22),
        ("Date / التاريخ", 14),
        ("Horaire / التوقيت", 15),
        ("Formateur / المدرب", 18),
        ("Matricule / رقم القيد", 16),
        ("Nom & Prénom / التلميذ", 28),
        ("Statut / الحالة", 15),
        ("Remarques / ملاحظات", 24),
    ]

    header_row = 4
    ws1.row_dimensions[header_row].height = 25

    for col_idx, (h_title, col_w) in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col_idx)
        cell.value = h_title
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        col_letter = get_column_letter(col_idx)
        ws1.column_dimensions[col_letter].width = col_w

    # 3. Remplissage des données
    current_row = 5
    for idx, att in enumerate(attendances, 1):
        ws1.row_dimensions[current_row].height = 20

        # Données de la séance
        session = att.session
        grp = session.group
        subj = grp.subject if grp else None

        subj_name = subj.get_bilingual_name() if subj else "Échecs"
        grp_name = grp.name_fr if grp else "Groupe"
        date_str = att.date.strftime("%d/%m/%Y")
        horaire_str = f"{session.start_time.strftime('%H:%M')} - {session.end_time.strftime('%H:%M')}"
        trainer_str = session.trainer_name_fr or "GCA Coach"

        st = att.student
        st_mat = st.registration_number if st else "-"
        st_name = st.get_bilingual_full_name() if st else "Inconnu"
        status_lbl = att.get_status_label(lang)

        # Style du statut
        status_bg = "FFFFFF"
        status_fg = "0F172A"
        if att.status == "present":
            status_bg = CLR_PRESENT
            status_fg = TXT_PRESENT
        elif att.status == "absent":
            status_bg = CLR_ABSENT
            status_fg = TXT_ABSENT
        elif att.status == "late":
            status_bg = CLR_LATE
            status_fg = TXT_LATE
        elif att.status == "justified":
            status_bg = CLR_JUSTIFIED
            status_fg = TXT_JUSTIFIED

        row_vals = [
            (idx, "center"),
            (subj_name, "left"),
            (grp_name, "left"),
            (date_str, "center"),
            (horaire_str, "center"),
            (trainer_str, "left"),
            (st_mat, "center"),
            (st_name, "left"),
            (status_lbl, "center"),
            (att.notes or "", "left"),
        ]

        for c_idx, (val, align_h) in enumerate(row_vals, 1):
            cell = ws1.cell(row=current_row, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9.5)
            cell.alignment = Alignment(horizontal=align_h, vertical="center")
            cell.border = cell_border

            # Alternance de fond léger
            if current_row % 2 == 0:
                cell.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")

            # Mettre en valeur la cellule Statut
            if c_idx == 9:
                cell.fill = PatternFill(start_color=status_bg, end_color=status_bg, fill_type="solid")
                cell.font = Font(name="Calibri", size=9.5, bold=True, color=status_fg)

        current_row += 1

    # =========================================================================
    # FEUILLE 2 : SYNTHÈSE D'ASSIDUITÉ PAR ÉLÈVE
    # =========================================================================
    ws2 = wb.create_sheet(title="Synthèse par Élève" if lang == "fr" else "ملخص الحضور لكل تلميذ")
    ws2.views.sheetView[0].showGridLines = True

    # 1. En-tête
    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value = (
        f"Synthèse d'Assiduité par Élève • {period_label}"
        if lang == "fr" else
        f"ملخص ونسبة المواظبة لكل تلميذ • {period_label}"
    )
    t2.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t2.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    headers2 = [
        ("N°", 6),
        ("Matricule / رقم القيد", 16),
        ("Nom & Prénom de l'Élève / التلميذ", 30),
        ("Total Séances / المجموع", 16),
        ("Présents / حاضر", 14),
        ("Retards / متأخر", 14),
        ("Absents / غائب", 14),
        ("Taux d'Assiduité / نسبة المواظبة", 22),
    ]

    ws2.row_dimensions[3].height = 24
    for c_idx, (h_title, col_w) in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=c_idx, value=h_title)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        col_letter = get_column_letter(c_idx)
        ws2.column_dimensions[col_letter].width = col_w

    row2 = 4
    for idx, s in enumerate(student_summaries, 1):
        ws2.row_dimensions[row2].height = 20
        rate_str = f"{s['rate']}%"

        # Couleur de fond du taux d'assiduité
        if s['rate'] >= 80:
            rate_bg = CLR_PRESENT
            rate_fg = TXT_PRESENT
        elif s['rate'] >= 50:
            rate_bg = CLR_LATE
            rate_fg = TXT_LATE
        else:
            rate_bg = CLR_ABSENT
            rate_fg = TXT_ABSENT

        vals2 = [
            (idx, "center"),
            (s['student'].registration_number, "center"),
            (s['student'].get_bilingual_full_name(), "left"),
            (s['total'], "center"),
            (s['present'], "center"),
            (s['late'], "center"),
            (s['absent'], "center"),
            (rate_str, "center"),
        ]

        for c_idx, (val, align_h) in enumerate(vals2, 1):
            cell = ws2.cell(row=row2, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=9.5)
            cell.alignment = Alignment(horizontal=align_h, vertical="center")
            cell.border = cell_border

            if row2 % 2 == 1:
                cell.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")

            if c_idx == 8:
                cell.fill = PatternFill(start_color=rate_bg, end_color=rate_bg, fill_type="solid")
                cell.font = Font(name="Calibri", size=9.5, bold=True, color=rate_fg)

        row2 += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

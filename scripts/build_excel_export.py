content = """import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY_FILL = PatternFill(start_color="001B57", end_color="001B57", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Segoe UI", size=14, bold=True, color="001B57")
DATA_FONT = Font(name="Segoe UI", size=10)
BORDER_THIN = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)

def export_students_to_excel(students_queryset, lang="fr"):
    \"\"\"
    Generates a bilingual Excel workbook (.xlsx) supporting full UTF-8,
    simultaneous French and Arabic characters (e.g. Mohamed العلوي),
    and RTL layout when lang == 'ar'.
    \"\"\"
    wb = Workbook()
    ws = wb.active

    if lang == "ar":
        ws.title = "قائمة التلاميذ"
        ws.sheet_view.rightToLeft = True
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="right", vertical="center")
        title_text = "أكاديمية جينيوس للشطرنج - لائحة التلاميذ المسجلين 2026"
        headers = [
            "رقم التسجيل", "الاسم بالعربية", "الاسم بالفرنسية",
            "النشاط والمستوى", "المجموعة", "ولي الأمر", "الهاتف", "الحالة"
        ]
    elif lang == "bilingual":
        ws.title = "Élèves - التلاميذ"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - Liste des Élèves / لائحة التلاميذ 2026"
        headers = [
            "Matricule / التسجيل", "Nom (FR)", "الاسم (AR)",
            "Activité / النشاط", "Groupe / المجموعة", "Parent / ولي الأمر", "Tél / الهاتف", "Statut / الحالة"
        ]
    else: # fr
        ws.title = "Liste des Élèves"
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_data = Alignment(horizontal="left", vertical="center")
        title_text = "GENIUS CHESS ACADEMY - Liste des Élèves Inscrits 2026"
        headers = [
            "Matricule", "Nom (Français)", "Nom (Arabe)",
            "Activité & Niveau", "Groupe", "Parent", "Téléphone", "Statut"
        ]

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers row
    ws.row_dimensions[3].height = 28
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = align_header
        cell.border = BORDER_THIN

    # Data rows
    row_idx = 4
    for st in students_queryset:
        ws.row_dimensions[row_idx].height = 22
        group = st.groups.first()
        subject_str = group.subject.get_bilingual_name() if group else "Échecs / الشطرنج"
        group_str = group.get_name(lang) if group else "-"
        parent_str = st.parent.get_name(lang) if st.parent else "-"
        phone_str = st.parent.phone if st.parent else "-"
        status_str = "Actif / نشط" if st.active else "Inactif / غير نشط"

        row_values = [
            st.registration_number,
            f"{st.first_name_ar} {st.last_name_ar}".strip(),
            f"{st.first_name_fr} {st.last_name_fr}".strip(),
            subject_str,
            group_str,
            parent_str,
            phone_str,
            status_str,
        ]
        if lang == "ar":
            # swap AR/FR column values to match AR headers
            row_values[1], row_values[2] = row_values[2], row_values[1]

        for col_num, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = DATA_FONT
            cell.alignment = align_data
            cell.border = BORDER_THIN
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    return excel_bytes
"""

with open('portal/excel_export.py', 'w', encoding='utf-8') as f:
    f.write(content)

print('Created portal/excel_export.py')
